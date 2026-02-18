from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Body, Depends
from fastapi.responses import FileResponse, StreamingResponse, Response
import cv2
import numpy as np
import json
import os
import uuid
from datetime import datetime, timedelta
import logging
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill
from ai_service import ai_service
import threading
import time
import pytz
import sys
from sqlalchemy.orm import Session
from database import get_db
import models

router = APIRouter(prefix="/api/productivity", tags=["productivity"])
logger = logging.getLogger(__name__)

# Store session data temporarily
sessions = {}

# Store active RTSP sessions
rtsp_sessions = {}

@router.post("/process")
async def process_video(
    video: UploadFile = File(...),
    model: str = Form(...),
    rois: str = Form(...)
):
    """Process uploaded video with ROI monitoring for factory productivity"""
    
    session_id = str(uuid.uuid4())
    temp_dir = f"uploads/productivity/{session_id}"
    os.makedirs(temp_dir, exist_ok=True)
    
    # Save uploaded video
    video_path = f"{temp_dir}/input.mp4"
    with open(video_path, "wb") as f:
        f.write(await video.read())
    
    # Parse ROIs
    roi_list = json.loads(rois)
    
    # Process video
    try:
        result = process_factory_video(
            video_path=video_path,
            model_name=model,
            rois=roi_list,
            output_dir=temp_dir
        )
        
        # Store session data
        sessions[session_id] = result
        
        return {
            "session_id": session_id,
            "processed_video_url": f"/uploads/productivity/{session_id}/output.mp4",
            "productivity_data": result
        }
    
    except Exception as e:
        logger.error(f"Error processing video: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download-excel")
async def download_excel(session_id: str):
    """Download Excel report for productivity data"""

    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    data = sessions[session_id]

    # Save to excel_reports folder
    excel_dir = "excel_reports"
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = f"{excel_dir}/productivity_report_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Generate Excel file
    generate_excel_report(data, excel_path)

    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"productivity_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@router.get("/download-excel-by-date")
async def download_excel_by_date(date: str):
    """Download Excel report for all sessions on a specific date (YYYY-MM-DD)"""

    try:
        target_date = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # Filter sessions by date
    ist = pytz.timezone('Asia/Kolkata')
    filtered_sessions = []

    for session_id, session_data in sessions.items():
        session_timestamp = session_data.get('timestamp', 0)
        session_datetime = datetime.fromtimestamp(session_timestamp, ist)

        if session_datetime.date() == target_date:
            filtered_sessions.append({
                'session_id': session_id,
                **session_data
            })

    if not filtered_sessions:
        raise HTTPException(status_code=404, detail=f"No sessions found for date {date}")

    # Generate combined Excel report
    excel_dir = "excel_reports"
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = f"{excel_dir}/productivity_report_{date}.xlsx"

    generate_daywise_excel_report(filtered_sessions, date, excel_path)

    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"productivity_report_{date}.xlsx"
    )


@router.get("/download-excel-filtered")
async def download_excel_filtered(
    start_date: str,
    end_date: str,
    camera: Optional[str] = None,
    report_type: str = "day-wise"
):
    """
    Download Excel report with date range and camera filters

    Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - camera: Camera/Machine number (1, 2, 3) or None for all cameras
    - report_type: 'day-wise' for separate sheets per day, 'combined' for single sheet
    """

    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if start > end:
        raise HTTPException(status_code=400, detail="Start date must be before or equal to end date")

    # Filter sessions by date range
    ist = pytz.timezone('Asia/Kolkata')
    filtered_sessions = []

    for session_id, session_data in sessions.items():
        session_timestamp = session_data.get('timestamp', 0)
        session_datetime = datetime.fromtimestamp(session_timestamp, ist)
        session_date = session_datetime.date()

        # Check if session is within date range
        if start <= session_date <= end:
            # Filter by camera if specified
            if camera and camera != 'all':
                # Filter ROI stats by machine number
                camera_num = int(camera)
                filtered_roi_stats = []

                for roi_stat in session_data.get('roi_stats', []):
                    # ROI IDs are 0-indexed, so machine 1 = ROI 0, machine 2 = ROI 1, etc.
                    if roi_stat.get('roi_id') == camera_num - 1:
                        filtered_roi_stats.append(roi_stat)

                if filtered_roi_stats:
                    # Create a copy of session data with filtered ROI stats
                    filtered_session = {
                        'session_id': session_id,
                        **session_data,
                        'roi_stats': filtered_roi_stats
                    }
                    filtered_sessions.append(filtered_session)
            else:
                # Include all ROI stats
                filtered_sessions.append({
                    'session_id': session_id,
                    **session_data
                })

    if not filtered_sessions:
        raise HTTPException(status_code=404, detail="No sessions found for selected filters")

    # Generate Excel based on report type
    excel_dir = "excel_reports"
    os.makedirs(excel_dir, exist_ok=True)

    if report_type == "day-wise":
        # Group sessions by date
        sessions_by_date = {}
        for session in filtered_sessions:
            session_timestamp = session.get('timestamp', 0)
            session_datetime = datetime.fromtimestamp(session_timestamp, ist)
            date_key = session_datetime.strftime('%Y-%m-%d')

            if date_key not in sessions_by_date:
                sessions_by_date[date_key] = []
            sessions_by_date[date_key].append(session)

        # Create Excel with separate sheets for each day
        excel_path = f"{excel_dir}/productivity_day_wise_{start_date}_to_{end_date}.xlsx"
        generate_multi_day_excel_report(sessions_by_date, excel_path)
        filename = f"productivity_day_wise_{start_date}_to_{end_date}.xlsx"
    else:
        # Combined report - all data in one sheet
        excel_path = f"{excel_dir}/productivity_combined_{start_date}_to_{end_date}.xlsx"
        generate_combined_excel_report(filtered_sessions, start_date, end_date, excel_path)
        filename = f"productivity_combined_{start_date}_to_{end_date}.xlsx"

    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


def process_factory_video(video_path: str, model_name: str, rois: List[List[Dict]], output_dir: str) -> Dict:
    """
    Process video with steel detection and ROI monitoring

    Returns productivity data with downtime tracking
    """

    # Load model
    model_path = f"models/{model_name}"
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model not found: {model_path}")

    # Open video
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError("Failed to open video file")

    fps = int(cap.get(cv2.CAP_PROP_FPS))
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # Output video writer
    output_path = f"{output_dir}/output.mp4"
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))

    # Get video start time in IST for accurate timestamp calculation
    ist = pytz.timezone('Asia/Kolkata')
    video_start_time = datetime.now(ist)

    # Initialize ROI tracking
    roi_trackers = []
    for i, roi in enumerate(rois):
        roi_trackers.append({
            'id': i,
            'roi': roi,
            'is_red': True,  # Start as red (no steel)
            'downtime_start': 0,  # Frame number when downtime started
            'downtime_periods': [],  # List of (start_time, end_time) tuples
            'total_downtime': 0,
            'productive_time': 0
        })

    frame_count = 0
    logger.info(f"Processing video: {total_frames} frames at {fps} FPS")
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        
        # Run detection
        detections = ai_service.detect_objects(frame, model_name, confidence=0.3)
        
        # Check each ROI for steel presence
        for tracker in roi_trackers:
            roi_points = np.array([[p['x'], p['y']] for p in tracker['roi']], dtype=np.int32)
            steel_detected = False
            
            # Check if any detection is inside this ROI
            for detection in detections:
                bbox = detection['bbox']
                center_x = (bbox[0] + bbox[2]) / 2
                center_y = (bbox[1] + bbox[3]) / 2
                
                # Check if center point is inside ROI polygon
                if cv2.pointPolygonTest(roi_points, (center_x, center_y), False) >= 0:
                    steel_detected = True
                    break
            
            # Update ROI state
            current_time = frame_count / fps
            
            if steel_detected:
                # Steel present - ROI should be green, TIMER STOPS
                if tracker['is_red']:
                    # Transition from red to green - end downtime period
                    downtime_end = current_time
                    downtime_start_time = tracker['downtime_start'] / fps
                    duration = downtime_end - downtime_start_time

                    # Calculate actual IST timestamps based on video start time
                    ist = pytz.timezone('Asia/Kolkata')
                    start_ist = video_start_time + timedelta(seconds=downtime_start_time)
                    end_ist = video_start_time + timedelta(seconds=downtime_end)

                    tracker['downtime_periods'].append({
                        'start': downtime_start_time,
                        'end': downtime_end,
                        'duration': duration,
                        'start_time_ist': start_ist.strftime('%H:%M:%S'),
                        'end_time_ist': end_ist.strftime('%H:%M:%S'),
                        'start_datetime_ist': start_ist.strftime('%Y-%m-%d %H:%M:%S'),
                        'end_datetime_ist': end_ist.strftime('%Y-%m-%d %H:%M:%S')
                    })
                    tracker['total_downtime'] += duration

                    # Console output: Timer stopped
                    logger.info(f"⏱️  Machine {tracker['id'] + 1}: TIMER STOPPED - Steel detected! Downtime: {duration:.2f}s ({start_ist.strftime('%H:%M:%S')} to {end_ist.strftime('%H:%M:%S')})")
                tracker['is_red'] = False
            else:
                # No steel - ROI should be red, TIMER STARTS
                if not tracker['is_red']:
                    # Transition from green to red - start downtime period
                    tracker['downtime_start'] = frame_count

                    # Console output: Timer started
                    logger.info(f"⏱️  Machine {tracker['id'] + 1}: TIMER STARTED - No steel detected at {current_time:.2f}s")
                tracker['is_red'] = True
        
        # Draw ROIs and timer on frame
        frame = draw_rois_and_timer(frame, roi_trackers, frame_count, fps)
        
        out.write(frame)
        frame_count += 1
        
        if frame_count % 100 == 0:
            logger.info(f"Processed {frame_count}/{total_frames} frames")
    
    cap.release()
    out.release()

    # Handle final downtime periods (if video ends while in red state)
    video_duration = total_frames / fps
    ist = pytz.timezone('Asia/Kolkata')

    for tracker in roi_trackers:
        if tracker['is_red']:
            downtime_start_time = tracker['downtime_start'] / fps
            duration = video_duration - downtime_start_time

            # Calculate IST timestamps
            start_ist = video_start_time + timedelta(seconds=downtime_start_time)
            end_ist = video_start_time + timedelta(seconds=video_duration)

            tracker['downtime_periods'].append({
                'start': downtime_start_time,
                'end': video_duration,
                'duration': duration,
                'start_time_ist': start_ist.strftime('%H:%M:%S'),
                'end_time_ist': end_ist.strftime('%H:%M:%S'),
                'start_datetime_ist': start_ist.strftime('%Y-%m-%d %H:%M:%S'),
                'end_datetime_ist': end_ist.strftime('%Y-%m-%d %H:%M:%S')
            })
            tracker['total_downtime'] += duration

        tracker['productive_time'] = video_duration - tracker['total_downtime']

    # Calculate statistics
    roi_stats = []
    for tracker in roi_trackers:
        efficiency = (tracker['productive_time'] / video_duration * 100) if video_duration > 0 else 0
        roi_stats.append({
            'roi_id': tracker['id'],
            'total_downtime': round(tracker['total_downtime'], 2),
            'productive_time': round(tracker['productive_time'], 2),
            'efficiency': round(efficiency, 2),
            'downtime_count': len(tracker['downtime_periods']),
            'downtime_periods': tracker['downtime_periods']
        })

    logger.info(f"Video processing complete: {frame_count} frames processed")

    result = {
        'session_id': output_dir.split('/')[-1],
        'video_duration': round(video_duration, 2),
        'total_frames': total_frames,
        'fps': fps,
        'roi_stats': roi_stats,
        'timestamp': time.time()
    }

    # Store in sessions for dashboard
    sessions[output_dir.split('/')[-1]] = result

    return result


def draw_rois_and_timer(frame: np.ndarray, roi_trackers: List[Dict], frame_count: int, fps: int,
                       detections: List[Dict] = None) -> np.ndarray:
    """Draw ROIs with color coding, timer, and steel bounding boxes with track IDs"""

    overlay = frame.copy()

    # Draw each ROI
    for tracker in roi_trackers:
        roi_points = np.array([[p['x'], p['y']] for p in tracker['roi']], dtype=np.int32)

        # Color based on state
        if tracker['is_red']:
            color = (0, 0, 255)  # Red (BGR)
            fill_color = (0, 0, 255, 80)  # Semi-transparent red
        else:
            color = (0, 255, 0)  # Green (BGR)
            fill_color = (0, 255, 0, 80)  # Semi-transparent green

        # Draw filled polygon
        cv2.fillPoly(overlay, [roi_points], color)

        # Draw border
        cv2.polylines(frame, [roi_points], True, color, 3)

        # Calculate center for label
        M = cv2.moments(roi_points)
        if M["m00"] != 0:
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])
        else:
            cx = roi_points[0][0]
            cy = roi_points[0][1]

        # Draw label with steel count
        label = f"Machine {tracker['id'] + 1}"
        status = "IDLE" if tracker['is_red'] else "ACTIVE"
        steel_count = tracker.get('steel_count', 0)

        cv2.putText(frame, label, (cx - 60, cy - 40),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
        cv2.putText(frame, status, (cx - 40, cy - 10),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        cv2.putText(frame, f"Count: {steel_count}", (cx - 50, cy + 20),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)

    # Blend overlay with original frame
    frame = cv2.addWeighted(overlay, 0.3, frame, 0.7, 0)

    # Draw steel bounding boxes with track IDs and confidence
    if detections:
        for detection in detections:
            if detection['class'] != 'steel':
                continue

            bbox = detection['bbox']
            x1, y1, x2, y2 = map(int, bbox)
            track_id = detection.get('track_id')
            confidence = detection.get('confidence', 0)

            # Check if this steel is being tracked (counted) in any ROI
            is_in_roi = False

            for tracker in roi_trackers:
                if track_id and track_id in tracker.get('tracked_steels', {}):
                    is_in_roi = True
                    break

            # Color coding:
            # Green = In ROI and counted
            # Yellow = Detected but not in ROI yet
            if is_in_roi:
                bbox_color = (0, 255, 0)  # Green - in ROI and counted
                label_text = f"ID:{track_id} | {confidence:.2f} | COUNTED"
            else:
                bbox_color = (0, 255, 255)  # Yellow - detected, not in ROI
                label_text = f"ID:{track_id} | {confidence:.2f}" if track_id else f"{confidence:.2f}"

            # Draw bounding box
            cv2.rectangle(frame, (x1, y1), (x2, y2), bbox_color, 2)

            # Draw label with background
            (text_width, text_height), _ = cv2.getTextSize(label_text, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 2)
            cv2.rectangle(frame, (x1, y1 - text_height - 10), (x1 + text_width + 10, y1), bbox_color, -1)
            cv2.putText(frame, label_text, (x1 + 5, y1 - 5),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)

    # Draw minimized timer at top-right corner with IST time
    ist = pytz.timezone('Asia/Kolkata')
    current_ist = datetime.now(ist)
    ist_time_text = current_ist.strftime('%H:%M:%S')

    # Minimized timer background (smaller size)
    cv2.rectangle(frame, (frame.shape[1] - 160, 10), (frame.shape[1] - 10, 45), (0, 0, 0), -1)
    cv2.rectangle(frame, (frame.shape[1] - 160, 10), (frame.shape[1] - 10, 45), (255, 255, 255), 1)

    # Draw IST time text (smaller font)
    cv2.putText(frame, ist_time_text, (frame.shape[1] - 150, 32),
               cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

    return frame


def generate_excel_report(data: Dict, output_path: str):
    """Generate Excel report with productivity data"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productivity Report"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Title
    ws['A1'] = "Factory Productivity Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    # Summary
    ws['A3'] = "Video Duration:"
    ws['B3'] = f"{data['video_duration']} seconds"
    ws['A4'] = "Total Frames:"
    ws['B4'] = data['total_frames']
    ws['A5'] = "FPS:"
    ws['B5'] = data['fps']
    ws['A6'] = "Generated:"
    ws['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Machine statistics
    row = 9
    for stat in data['roi_stats']:
        ws[f'A{row}'] = f"Machine {stat['roi_id'] + 1} Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Stats table
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1

        ws[f'A{row}'] = "Total Downtime"
        ws[f'B{row}'] = f"{stat['total_downtime']} seconds"
        row += 1

        ws[f'A{row}'] = "Productive Time"
        ws[f'B{row}'] = f"{stat['productive_time']} seconds"
        row += 1

        ws[f'A{row}'] = "Efficiency"
        ws[f'B{row}'] = f"{stat['efficiency']}%"
        row += 1

        ws[f'A{row}'] = "Downtime Events"
        ws[f'B{row}'] = stat['downtime_count']
        row += 2

        # Downtime periods table
        if stat['downtime_periods']:
            ws[f'A{row}'] = "Downtime Periods"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Event #"
            ws[f'B{row}'] = "Start Time (IST)"
            ws[f'C{row}'] = "End Time (IST)"
            ws[f'D{row}'] = "Duration (s)"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].font = header_font
            row += 1

            for i, period in enumerate(stat['downtime_periods'], 1):
                ws[f'A{row}'] = i
                ws[f'B{row}'] = period.get('start_time_ist', f"{round(period['start'], 2)}s")
                ws[f'C{row}'] = period.get('end_time_ist', f"{round(period['end'], 2)}s")
                ws[f'D{row}'] = round(period['duration'], 2)
                row += 1

        row += 2

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    logger.info(f"Excel report generated: {output_path}")


def format_duration(seconds):
    """Format seconds to minutes or hours (0-60 minutes, then convert to hours)"""
    if not seconds or seconds == 0:
        return "0 min"

    total_minutes = int(seconds // 60)
    remaining_seconds = int(seconds % 60)

    if total_minutes >= 60:
        hours = total_minutes // 60
        minutes = total_minutes % 60
        if minutes > 0:
            return f"{hours} hr {minutes} min"
        return f"{hours} hr"

    if total_minutes > 0 and remaining_seconds > 0:
        return f"{total_minutes} min {remaining_seconds} sec"
    elif total_minutes > 0:
        return f"{total_minutes} min"
    else:
        return f"{remaining_seconds} sec"


def generate_daywise_excel_report(sessions_list: List[Dict], date: str, output_path: str):
    """Generate Excel report for all sessions on a specific day"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {date}"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)

    # Title
    ws['A1'] = f"Factory Productivity Report - {date}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')

    # Summary
    total_sessions = len(sessions_list)
    total_downtime = sum(sum(stat['total_downtime'] for stat in session.get('roi_stats', [])) for session in sessions_list)
    total_productive = sum(sum(stat['productive_time'] for stat in session.get('roi_stats', [])) for session in sessions_list)

    ws['A3'] = "Date:"
    ws['B3'] = date
    ws['A4'] = "Total Sessions:"
    ws['B4'] = total_sessions
    ws['A5'] = "Total Downtime:"
    ws['B5'] = format_duration(total_downtime)
    ws['A6'] = "Total Productive Time:"
    ws['B6'] = format_duration(total_productive)
    ws['A7'] = "Generated:"
    ws['B7'] = datetime.now(pytz.timezone('Asia/Kolkata')).strftime("%Y-%m-%d %H:%M:%S IST")

    row = 10

    # Process each session
    for session_idx, session in enumerate(sessions_list, 1):
        ws[f'A{row}'] = f"Session {session_idx}"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="0000FF")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws[f'A{row}'] = "Session ID:"
        ws[f'B{row}'] = session.get('session_id', 'N/A')
        row += 1

        ws[f'A{row}'] = "Duration:"
        ws[f'B{row}'] = format_duration(session.get('video_duration', 0))
        row += 1

        # Machine statistics for this session
        for stat in session.get('roi_stats', []):
            ws[f'A{row}'] = f"  Machine {stat['roi_id'] + 1}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            # Only show downtime and productive time
            ws[f'A{row}'] = "    Downtime:"
            ws[f'B{row}'] = format_duration(stat['total_downtime'])
            ws[f'B{row}'].font = Font(color="FF0000")
            row += 1

            ws[f'A{row}'] = "    Productive Time:"
            ws[f'B{row}'] = format_duration(stat['productive_time'])
            ws[f'B{row}'].font = Font(color="00FF00")
            row += 1

            # Downtime periods with IST timestamps
            if stat.get('downtime_periods'):
                ws[f'A{row}'] = "    Downtime Periods:"
                ws[f'A{row}'].font = Font(bold=True, italic=True)
                row += 1

                # Headers
                ws[f'A{row}'] = "      Event"
                ws[f'B{row}'] = "Start Time (IST)"
                ws[f'C{row}'] = "End Time (IST)"
                ws[f'D{row}'] = "Duration"
                ws[f'E{row}'] = "Video Recording"
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = header_fill
                    ws[f'{col}{row}'].font = header_font
                row += 1

                for i, period in enumerate(stat['downtime_periods'], 1):
                    ws[f'A{row}'] = f"      {i}"

                    # Debug: Log the entire period dictionary
                    logger.info(f"📊 Period {i} data: {period}")

                    # Use IST time format (HH:MM:SS) - calculate if not present
                    if 'start_time_ist' in period and 'end_time_ist' in period:
                        start_time_ist = period['start_time_ist']
                        end_time_ist = period['end_time_ist']
                    else:
                        # Fallback: Calculate IST from session timestamp + elapsed seconds
                        session_timestamp = session.get('timestamp', time.time())
                        ist = pytz.timezone('Asia/Kolkata')
                        session_start = datetime.fromtimestamp(session_timestamp, ist)

                        start_seconds = period.get('start', 0)
                        end_seconds = period.get('end', 0)

                        start_time = session_start + timedelta(seconds=start_seconds)
                        end_time = session_start + timedelta(seconds=end_seconds)

                        start_time_ist = start_time.strftime('%H:%M:%S')
                        end_time_ist = end_time.strftime('%H:%M:%S')

                    logger.info(f"📊 Period {i} - start_time_ist: {start_time_ist}, end_time_ist: {end_time_ist}")

                    ws[f'B{row}'] = start_time_ist  # Format: 10:20:20
                    ws[f'C{row}'] = end_time_ist    # Format: 10:12:40
                    ws[f'D{row}'] = format_duration(period['duration'])

                    # Add video link if recording exists
                    video_path = period.get('video_path', '')

                    # Debug logging
                    logger.info(f"📹 Excel Gen - Checking video_path: {video_path}")
                    logger.info(f"📹 Excel Gen - File exists: {os.path.exists(video_path) if video_path else False}")

                    if video_path and os.path.exists(video_path):
                        # Get absolute path for local file access
                        absolute_video_path = os.path.abspath(video_path)

                        # Use Excel HYPERLINK formula with local file path (clickable to open video)
                        ws[f'E{row}'] = f'=HYPERLINK("{absolute_video_path}", "📹 {start_time_ist}")'
                        ws[f'E{row}'].font = Font(color="0000FF", underline="single")
                        logger.info(f"✅ Excel Gen - Added video link: {absolute_video_path}")
                    else:
                        ws[f'E{row}'] = "No recording"
                        ws[f'E{row}'].font = Font(color="999999")
                        logger.info(f"❌ Excel Gen - No recording found for period {i}")

                    row += 1

            row += 1

        row += 2

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25

    wb.save(output_path)
    logger.info(f"Day-wise Excel report generated: {output_path}")


def generate_multi_day_excel_report(sessions_by_date: Dict[str, List[Dict]], output_path: str):
    """Generate Excel report with separate sheets for each day"""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create a sheet for each date
    for date, sessions_list in sorted(sessions_by_date.items()):
        ws = wb.create_sheet(title=date)

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)

        # Title
        ws['A1'] = f"Productivity Report - {date}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')

        # Summary
        total_sessions = len(sessions_list)
        total_downtime = sum(sum(stat['total_downtime'] for stat in session.get('roi_stats', [])) for session in sessions_list)
        total_productive = sum(sum(stat['productive_time'] for stat in session.get('roi_stats', [])) for session in sessions_list)

        ws['A3'] = "Date:"
        ws['B3'] = date
        ws['A4'] = "Total Sessions:"
        ws['B4'] = total_sessions
        ws['A5'] = "Total Downtime:"
        ws['B5'] = format_duration(total_downtime)
        ws['A6'] = "Total Productive Time:"
        ws['B6'] = format_duration(total_productive)

        row = 9

        # Process each session
        for session_idx, session in enumerate(sessions_list, 1):
            ws[f'A{row}'] = f"Session {session_idx}"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="0000FF")
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            # Machine statistics for this session
            for stat in session.get('roi_stats', []):
                ws[f'A{row}'] = f"Machine {stat['roi_id'] + 1}"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws.merge_cells(f'A{row}:E{row}')
                row += 1

                # Stats summary
                ws[f'A{row}'] = "Downtime:"
                ws[f'B{row}'] = format_duration(stat['total_downtime'])
                ws[f'C{row}'] = "Productive:"
                ws[f'D{row}'] = format_duration(stat['productive_time'])
                row += 1

                # Downtime periods header
                if stat.get('downtime_periods'):
                    ws[f'A{row}'] = "Event"
                    ws[f'B{row}'] = "Start Time"
                    ws[f'C{row}'] = "End Time"
                    ws[f'D{row}'] = "Duration"
                    ws[f'E{row}'] = "Video Recording"

                    for col in ['A', 'B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].fill = header_fill
                        ws[f'{col}{row}'].font = header_font
                    row += 1

                    # Downtime periods
                    for i, period in enumerate(stat['downtime_periods'], 1):
                        ws[f'A{row}'] = f"      {i}"

                        # Calculate IST timestamps if not present
                        if 'start_time_ist' in period and 'end_time_ist' in period:
                            start_time_ist = period['start_time_ist']
                            end_time_ist = period['end_time_ist']
                        else:
                            # Fallback: Calculate from session timestamp
                            session_timestamp = session.get('timestamp', time.time())
                            ist = pytz.timezone('Asia/Kolkata')
                            session_start = datetime.fromtimestamp(session_timestamp, ist)

                            start_time = session_start + timedelta(seconds=period.get('start', 0))
                            end_time = session_start + timedelta(seconds=period.get('end', 0))

                            start_time_ist = start_time.strftime('%H:%M:%S')
                            end_time_ist = end_time.strftime('%H:%M:%S')

                        ws[f'B{row}'] = start_time_ist
                        ws[f'C{row}'] = end_time_ist
                        ws[f'D{row}'] = format_duration(period['duration'])

                        # Add video link
                        video_path = period.get('video_path', '')
                        if video_path and os.path.exists(video_path):
                            # Get absolute path for local file access
                            absolute_video_path = os.path.abspath(video_path)

                            # Use Excel HYPERLINK formula with local file path (clickable to open video)
                            ws[f'E{row}'] = f'=HYPERLINK("{absolute_video_path}", "📹 {start_time_ist}")'
                            ws[f'E{row}'].font = Font(color="0000FF", underline="single")
                        else:
                            ws[f'E{row}'] = "No recording"

                        row += 1

                row += 1

            row += 2

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25

    wb.save(output_path)
    logger.info(f"Multi-day Excel report generated: {output_path}")


def generate_combined_excel_report(sessions_list: List[Dict], start_date: str, end_date: str, output_path: str):
    """Generate combined Excel report with all sessions in a single sheet"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Report"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)

    # Title
    ws['A1'] = f"Combined Productivity Report ({start_date} to {end_date})"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')

    # Summary
    total_sessions = len(sessions_list)
    total_downtime = sum(sum(stat['total_downtime'] for stat in session.get('roi_stats', [])) for session in sessions_list)
    total_productive = sum(sum(stat['productive_time'] for stat in session.get('roi_stats', [])) for session in sessions_list)

    ws['A3'] = "Date Range:"
    ws['B3'] = f"{start_date} to {end_date}"
    ws['A4'] = "Total Sessions:"
    ws['B4'] = total_sessions
    ws['A5'] = "Total Downtime:"
    ws['B5'] = format_duration(total_downtime)
    ws['A6'] = "Total Productive Time:"
    ws['B6'] = format_duration(total_productive)
    ws['A7'] = "Generated:"
    ws['B7'] = datetime.now(pytz.timezone('Asia/Kolkata')).strftime("%Y-%m-%d %H:%M:%S IST")

    row = 10

    # Process each session
    for session_idx, session in enumerate(sessions_list, 1):
        # Get session date
        session_timestamp = session.get('timestamp', 0)
        ist = pytz.timezone('Asia/Kolkata')
        session_datetime = datetime.fromtimestamp(session_timestamp, ist)
        session_date = session_datetime.strftime('%Y-%m-%d')

        ws[f'A{row}'] = f"Session {session_idx} - {session_date}"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="0000FF")
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        ws[f'A{row}'] = "Session ID:"
        ws[f'B{row}'] = session.get('session_id', 'N/A')
        row += 1

        # Machine statistics for this session
        for stat in session.get('roi_stats', []):
            ws[f'A{row}'] = f"Machine {stat['roi_id'] + 1}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            # Stats summary
            ws[f'A{row}'] = "Downtime:"
            ws[f'B{row}'] = format_duration(stat['total_downtime'])
            ws[f'C{row}'] = "Productive:"
            ws[f'D{row}'] = format_duration(stat['productive_time'])
            row += 1

            # Downtime periods header
            if stat.get('downtime_periods'):
                ws[f'A{row}'] = "Event"
                ws[f'B{row}'] = "Start Time"
                ws[f'C{row}'] = "End Time"
                ws[f'D{row}'] = "Duration"
                ws[f'E{row}'] = "Video Recording"

                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = header_fill
                    ws[f'{col}{row}'].font = header_font
                row += 1

                # Downtime periods
                for i, period in enumerate(stat['downtime_periods'], 1):
                    ws[f'A{row}'] = f"      {i}"

                    # Calculate IST timestamps if not present
                    if 'start_time_ist' in period and 'end_time_ist' in period:
                        start_time_ist = period['start_time_ist']
                        end_time_ist = period['end_time_ist']
                    else:
                        # Fallback: Calculate from session timestamp
                        session_timestamp = session.get('timestamp', time.time())
                        ist = pytz.timezone('Asia/Kolkata')
                        session_start = datetime.fromtimestamp(session_timestamp, ist)

                        start_time = session_start + timedelta(seconds=period.get('start', 0))
                        end_time = session_start + timedelta(seconds=period.get('end', 0))

                        start_time_ist = start_time.strftime('%H:%M:%S')
                        end_time_ist = end_time.strftime('%H:%M:%S')

                    ws[f'B{row}'] = start_time_ist
                    ws[f'C{row}'] = end_time_ist
                    ws[f'D{row}'] = format_duration(period['duration'])

                    # Add video link
                    video_path = period.get('video_path', '')
                    if video_path and os.path.exists(video_path):
                        # Get absolute path for local file access
                        absolute_video_path = os.path.abspath(video_path)

                        # Use Excel HYPERLINK formula with local file path (clickable to open video)
                        ws[f'E{row}'] = f'=HYPERLINK("{absolute_video_path}", "📹 {start_time_ist}")'
                        ws[f'E{row}'].font = Font(color="0000FF", underline="single")
                    else:
                        ws[f'E{row}'] = "No recording"

                    row += 1

            row += 1

        row += 2

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25

    wb.save(output_path)
    logger.info(f"Combined Excel report generated: {output_path}")


# ==================== RTSP ENDPOINTS ====================

@router.post("/rtsp/connect")
async def connect_to_rtsp(data: Dict = Body(...)):
    """Connect to RTSP stream and get first frame for ROI drawing"""

    rtsp_url = data.get('rtsp_url')

    if not rtsp_url:
        raise HTTPException(status_code=400, detail="RTSP URL is required")

    try:
        # Open RTSP stream with optimized settings
        cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)

        # Set timeouts and buffer settings
        cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 10000)  # 10 second timeout for connection
        cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 10000)  # 10 second timeout for reading
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

        if not cap.isOpened():
            raise HTTPException(status_code=400, detail="Failed to connect to RTSP stream. Check URL and network connection.")

        # Try to read first frame (may need multiple attempts due to buffering)
        ret = False
        frame = None
        max_attempts = 10

        for attempt in range(max_attempts):
            ret, frame = cap.read()
            if ret:
                break
            time.sleep(0.5)

        cap.release()

        if not ret or frame is None:
            raise HTTPException(status_code=400, detail="Failed to read frame from RTSP stream. Stream may be unavailable.")

        # Save frame temporarily
        temp_id = str(uuid.uuid4())
        temp_dir = "uploads/productivity/temp"
        os.makedirs(temp_dir, exist_ok=True)
        frame_path = f"{temp_dir}/{temp_id}.jpg"
        cv2.imwrite(frame_path, frame)

        return {
            "message": "Connected to RTSP stream",
            "frame_url": f"/api/productivity/temp-frame/{temp_id}"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error connecting to RTSP: {str(e)}")


@router.get("/temp-frame/{frame_id}")
async def get_temp_frame(frame_id: str):
    """Get temporary frame for ROI drawing"""

    frame_path = f"uploads/productivity/temp/{frame_id}.jpg"

    if not os.path.exists(frame_path):
        raise HTTPException(status_code=404, detail="Frame not found")

    return FileResponse(frame_path, media_type="image/jpeg")


@router.post("/rtsp/start")
async def start_rtsp_monitoring(data: Dict = Body(...)):
    """Start RTSP stream monitoring with ROI tracking"""

    rtsp_url = data.get('rtsp_url')
    model_name = data.get('model')
    rois = data.get('rois')

    if not rtsp_url or not model_name or not rois:
        raise HTTPException(status_code=400, detail="Missing required parameters")

    # Check for existing active sessions with same RTSP URL
    for existing_id, existing_session in list(rtsp_sessions.items()):
        if (existing_session.get('rtsp_url') == rtsp_url and
            existing_session.get('is_active') and
            not existing_session.get('stop_requested')):
            # Return existing session instead of creating duplicate
            logger.info(f"Returning existing session: {existing_id}")
            return {
                "session_id": existing_id,
                "message": "Using existing RTSP monitoring session"
            }

    session_id = str(uuid.uuid4())

    # Initialize session data
    rtsp_sessions[session_id] = {
        'rtsp_url': rtsp_url,
        'model_name': model_name,
        'rois': rois,
        'is_active': True,
        'start_time': time.time(),
        'frame_count': 0,
        'latest_frame': None,
        'roi_trackers': [],
        'stop_requested': False
    }

    # Initialize ROI trackers
    for i, roi in enumerate(rois):
        rtsp_sessions[session_id]['roi_trackers'].append({
            'id': i,
            'roi': roi,
            'is_red': True,  # Start as red (no steel)
            'downtime_start': 0,
            'downtime_periods': [],
            'total_downtime': 0,
            'productive_time': 0,
            'current_downtime': 0
        })

    # Start RTSP processing thread
    thread = threading.Thread(
        target=process_rtsp_stream,
        args=(session_id,),
        daemon=True
    )
    thread.start()

    logger.info(f"Started RTSP monitoring session: {session_id}")

    return {
        "session_id": session_id,
        "message": "RTSP monitoring started"
    }


@router.post("/rtsp/stop/{session_id}")
async def stop_rtsp_monitoring(session_id: str, db: Session = Depends(get_db)):
    """Stop RTSP stream monitoring and generate report"""

    if session_id not in rtsp_sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = rtsp_sessions[session_id]

    # Mark session as stopped
    session['stop_requested'] = True
    session['is_active'] = False

    # Wait for thread to finish processing
    time.sleep(1.5)

    # Calculate final statistics
    elapsed_time = time.time() - session['start_time']

    for tracker in session['roi_trackers']:
        # Handle final downtime period if still red
        if tracker['is_red']:
            downtime_duration = elapsed_time - tracker['downtime_start']
            tracker['downtime_periods'].append({
                'start': tracker['downtime_start'],
                'end': elapsed_time,
                'duration': downtime_duration
            })
            tracker['total_downtime'] += downtime_duration

        tracker['productive_time'] = elapsed_time - tracker['total_downtime']

    # Generate productivity data
    roi_stats = []
    for tracker in session['roi_trackers']:
        efficiency = (tracker['productive_time'] / elapsed_time * 100) if elapsed_time > 0 else 0
        roi_stats.append({
            'roi_id': tracker['id'],
            'total_downtime': round(tracker['total_downtime'], 2),
            'productive_time': round(tracker['productive_time'], 2),
            'efficiency': round(efficiency, 2),
            'downtime_count': len(tracker['downtime_periods']),
            'downtime_periods': tracker['downtime_periods'],
            'steel_count': tracker.get('steel_count', 0)  # Add steel count
        })

    productivity_data = {
        'session_id': session_id,
        'video_duration': round(elapsed_time, 2),
        'total_frames': session['frame_count'],
        'fps': round(session['frame_count'] / elapsed_time, 2) if elapsed_time > 0 else 0,
        'roi_stats': roi_stats,
        'timestamp': time.time()
    }

    # Store in sessions for Excel download and dashboard
    sessions[session_id] = productivity_data

    # ✅ SAVE TO DATABASE FOR PERSISTENCE
    try:
        # Check if session already exists in database
        db_session = db.query(models.ProductivitySession).filter(
            models.ProductivitySession.session_id == session_id
        ).first()

        if db_session:
            # Update existing session
            logger.info(f"📝 Updating existing session in database: {session_id}")
            db_session.end_time = datetime.fromtimestamp(time.time())
            db_session.video_duration = round(elapsed_time, 2)
            db_session.total_frames = session['frame_count']
            db_session.fps = round(session['frame_count'] / elapsed_time, 2) if elapsed_time > 0 else 0
            db_session.roi_stats = roi_stats
            db_session.video_recordings = session.get('video_recordings', [])
            db_session.updated_at = datetime.now()
        else:
            # Create new session record
            logger.info(f"➕ Creating new session in database: {session_id}")
            db_session = models.ProductivitySession(
                session_id=session_id,
                rtsp_url=session.get('rtsp_url'),
                model_name=session.get('model_name'),
                start_time=datetime.fromtimestamp(session['start_time']),
                end_time=datetime.fromtimestamp(time.time()),
                video_duration=round(elapsed_time, 2),
                total_frames=session['frame_count'],
                fps=round(session['frame_count'] / elapsed_time, 2) if elapsed_time > 0 else 0,
                roi_stats=roi_stats,
                video_recordings=session.get('video_recordings', [])
            )
            db.add(db_session)

        db.commit()
        db.refresh(db_session)  # Refresh to get the latest data
        logger.info(f"✅ Successfully saved productivity session to database: {session_id}")
        logger.info(f"📊 Session data: duration={db_session.video_duration}s, frames={db_session.total_frames}, fps={db_session.fps}")
    except Exception as e:
        logger.error(f"❌ Failed to save session to database: {e}")
        logger.exception(e)  # Print full stack trace
        db.rollback()

    # Generate Excel report
    excel_dir = "excel_reports"
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = f"{excel_dir}/productivity_report_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    generate_excel_report(productivity_data, excel_path)

    logger.info(f"Stopped RTSP monitoring session: {session_id}")

    return {
        "message": "RTSP monitoring stopped",
        "productivity_data": productivity_data
    }


@router.get("/rtsp/stats/{session_id}")
async def get_rtsp_stats(session_id: str):
    """Get live statistics for active RTSP session"""

    if session_id not in rtsp_sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = rtsp_sessions[session_id]
    elapsed_time = time.time() - session['start_time']

    roi_stats = []
    for tracker in session['roi_trackers']:
        # Calculate current downtime if red
        current_downtime = 0
        if tracker['is_red']:
            current_downtime = elapsed_time - tracker['downtime_start']

        productive_time = elapsed_time - (tracker['total_downtime'] + current_downtime)

        roi_stats.append({
            'roi_id': tracker['id'],
            'is_red': tracker['is_red'],
            'current_downtime': round(current_downtime, 2),
            'total_downtime': round(tracker['total_downtime'], 2),
            'productive_time': round(productive_time, 2),
            'downtime_count': len(tracker['downtime_periods']),
            'steel_count': tracker.get('steel_count', 0)  # Add steel count
        })

    return {
        'elapsed_time': round(elapsed_time, 2),
        'frame_count': session['frame_count'],
        'roi_stats': roi_stats
    }


@router.get("/dashboard/stats")
async def get_dashboard_stats(db: Session = Depends(get_db)):
    """Get productivity dashboard statistics - active sessions and historical data from database"""

    # Get active RTSP sessions
    active_sessions = []
    for session_id, session in list(rtsp_sessions.items()):
        if session.get('is_active') and not session.get('stop_requested'):
            elapsed_time = time.time() - session['start_time']

            roi_stats = []
            for tracker in session['roi_trackers']:
                # Calculate current downtime if red
                current_downtime = 0
                if tracker['is_red']:
                    current_downtime = elapsed_time - tracker['downtime_start']

                productive_time = elapsed_time - (tracker['total_downtime'] + current_downtime)

                roi_stats.append({
                    'roi_id': tracker['id'],
                    'is_red': tracker['is_red'],
                    'current_downtime': round(current_downtime, 2),
                    'total_downtime': round(tracker['total_downtime'], 2),
                    'productive_time': round(productive_time, 2),
                    'downtime_count': len(tracker['downtime_periods']),
                    'steel_count': tracker.get('steel_count', 0)  # Add steel count
                })

            active_sessions.append({
                'session_id': session_id,
                'elapsed_time': round(elapsed_time, 2),
                'frame_count': session['frame_count'],
                'roi_stats': roi_stats,
                'timestamp': session['start_time']
            })

    # ✅ GET HISTORICAL SESSIONS FROM DATABASE (PERSISTENT STORAGE)
    historical_sessions = []
    try:
        # Query database for ALL completed sessions (most recent first)
        db_sessions = db.query(models.ProductivitySession).order_by(
            models.ProductivitySession.created_at.desc()
        ).all()

        logger.info(f"📊 Found {len(db_sessions)} total sessions in database")

        for db_session in db_sessions:
            # Skip if session is currently active
            if db_session.session_id in rtsp_sessions and rtsp_sessions[db_session.session_id].get('is_active'):
                logger.info(f"⏭️ Skipping active session: {db_session.session_id}")
                continue

            historical_sessions.append({
                'session_id': db_session.session_id,
                'video_duration': db_session.video_duration,
                'total_frames': db_session.total_frames,
                'fps': db_session.fps,
                'roi_stats': db_session.roi_stats or [],
                'timestamp': db_session.start_time.timestamp() if db_session.start_time else time.time()
            })

        logger.info(f"✅ Loaded {len(historical_sessions)} historical sessions from database")
    except Exception as e:
        logger.error(f"❌ Failed to load sessions from database: {e}")
        logger.exception(e)  # Print full stack trace
        # Fallback to in-memory sessions if database fails
        for session_id, session_data in sessions.items():
            if session_id not in rtsp_sessions or not rtsp_sessions[session_id].get('is_active'):
                historical_sessions.append({
                    'session_id': session_id,
                    'video_duration': session_data.get('video_duration', 0),
                    'total_frames': session_data.get('total_frames', 0),
                    'fps': session_data.get('fps', 0),
                    'roi_stats': session_data.get('roi_stats', []),
                    'timestamp': session_data.get('timestamp', time.time())
                })

    # Sort historical sessions by timestamp (most recent first)
    historical_sessions.sort(key=lambda x: x.get('timestamp', 0), reverse=True)

    logger.info(f"📤 Returning {len(active_sessions)} active + {len(historical_sessions)} historical sessions")

    return {
        'active_sessions': active_sessions,
        'historical_sessions': historical_sessions  # Return ALL sessions (no limit)
    }


@router.delete("/session/{session_id}")
async def delete_session(session_id: str, db: Session = Depends(get_db)):
    """Delete an active or historical session from memory and database"""

    # Check if it's an active RTSP session
    if session_id in rtsp_sessions:
        session = rtsp_sessions[session_id]
        session['stop_requested'] = True
        session['is_active'] = False

        # Wait a moment for thread to finish
        time.sleep(0.5)

        # Remove from active sessions
        del rtsp_sessions[session_id]

    # Remove from historical sessions (in-memory)
    if session_id in sessions:
        del sessions[session_id]

    # ✅ DELETE FROM DATABASE
    try:
        db_session = db.query(models.ProductivitySession).filter(
            models.ProductivitySession.session_id == session_id
        ).first()

        if db_session:
            db.delete(db_session)
            db.commit()
            logger.info(f"✅ Deleted session from database: {session_id}")
    except Exception as e:
        logger.error(f"❌ Failed to delete session from database: {e}")
        db.rollback()

    return {"message": "Session deleted successfully"}


@router.get("/recordings/{filename}")
async def get_recording(filename: str):
    """Serve recorded video files"""

    video_path = os.path.join("recordings", filename)

    if not os.path.exists(video_path):
        raise HTTPException(status_code=404, detail="Recording not found")

    return FileResponse(
        video_path,
        media_type="video/mp4",
        filename=filename
    )


@router.get("/video-player/{filename}")
async def video_player(filename: str, t: float = 0):
    """Serve video player page with timestamp support"""

    video_path = os.path.join("recordings", filename)

    if not os.path.exists(video_path):
        raise HTTPException(status_code=404, detail="Recording not found")

    # Return HTML page with video player that seeks to timestamp
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Video Player - {filename}</title>
        <style>
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            body {{
                background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                min-height: 100vh;
                padding: 20px;
            }}
            .container {{
                max-width: 1200px;
                width: 100%;
                background: rgba(255, 255, 255, 0.05);
                backdrop-filter: blur(10px);
                border-radius: 20px;
                padding: 30px;
                box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
                border: 1px solid rgba(255, 255, 255, 0.1);
            }}
            h1 {{
                color: #fff;
                margin-bottom: 10px;
                font-size: 28px;
                text-align: center;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                background-clip: text;
            }}
            .info {{
                color: #a0a0a0;
                text-align: center;
                margin-bottom: 20px;
                font-size: 14px;
            }}
            .video-wrapper {{
                position: relative;
                width: 100%;
                background: #000;
                border-radius: 15px;
                overflow: hidden;
                box-shadow: 0 10px 40px rgba(0, 0, 0, 0.5);
            }}
            video {{
                width: 100%;
                height: auto;
                display: block;
            }}
            .controls {{
                margin-top: 20px;
                display: flex;
                gap: 15px;
                justify-content: center;
                flex-wrap: wrap;
            }}
            .btn {{
                padding: 12px 24px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border: none;
                border-radius: 10px;
                cursor: pointer;
                font-size: 14px;
                font-weight: 600;
                transition: all 0.3s ease;
                box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
            }}
            .btn:hover {{
                transform: translateY(-2px);
                box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
            }}
            .btn:active {{
                transform: translateY(0);
            }}
            .timestamp {{
                color: #fff;
                background: rgba(255, 255, 255, 0.1);
                padding: 12px 20px;
                border-radius: 10px;
                font-size: 14px;
                display: flex;
                align-items: center;
                gap: 8px;
            }}
            .timestamp span {{
                color: #667eea;
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📹 Downtime Recording</h1>
            <div class="info">
                <p>{filename}</p>
                {f'<p style="color: #667eea; font-weight: bold; margin-top: 5px;">Starting at {t:.1f}s</p>' if t > 0 else ''}
            </div>
            <div class="video-wrapper">
                <video id="video" controls autoplay>
                    <source src="/api/productivity/recordings/{filename}" type="video/mp4">
                    Your browser does not support the video tag.
                </video>
            </div>
            <div class="controls">
                <div class="timestamp">
                    <span>Current Time:</span>
                    <span id="currentTime">0:00</span>
                </div>
                <button class="btn" onclick="video.currentTime = 0">⏮️ Restart</button>
                <button class="btn" onclick="video.playbackRate = 0.5">🐌 0.5x</button>
                <button class="btn" onclick="video.playbackRate = 1.0">▶️ 1x</button>
                <button class="btn" onclick="video.playbackRate = 2.0">⚡ 2x</button>
            </div>
        </div>
        <script>
            const video = document.getElementById('video');
            const currentTimeDisplay = document.getElementById('currentTime');

            // Seek to timestamp when video is loaded
            video.addEventListener('loadedmetadata', function() {{
                video.currentTime = {t};
            }});

            // Update current time display
            video.addEventListener('timeupdate', function() {{
                const minutes = Math.floor(video.currentTime / 60);
                const seconds = Math.floor(video.currentTime % 60);
                currentTimeDisplay.textContent = minutes + ':' + (seconds < 10 ? '0' : '') + seconds;
            }});

            // Auto-play after seeking
            video.addEventListener('seeked', function() {{
                video.play();
            }});
        </script>
    </body>
    </html>
    """

    return Response(content=html_content, media_type="text/html")


@router.get("/rtsp/feed/{session_id}")
async def get_rtsp_feed(session_id: str):
    """Stream RTSP feed with ROI overlay - 60 FPS for ultra-smooth playback"""

    if session_id not in rtsp_sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    def generate_frames():
        """Generate video frames at 60 FPS for ultra-smooth playback"""
        import time
        last_frame_time = time.time()

        # Target 60 FPS for smooth playback
        target_fps = 60
        min_frame_interval = 1.0 / target_fps

        while session_id in rtsp_sessions and rtsp_sessions[session_id]['is_active']:
            current_time = time.time()

            # Minimal frame rate limiting for 60 FPS
            elapsed = current_time - last_frame_time
            if elapsed < min_frame_interval:
                time.sleep(min_frame_interval - elapsed)

            session = rtsp_sessions[session_id]
            frame = session.get('latest_frame')

            if frame is None:
                time.sleep(0.001)  # 1ms delay (same as Live Monitoring)
                continue

            # Fast JPEG encoding - balanced quality (same as Live Monitoring)
            encode_params = [cv2.IMWRITE_JPEG_QUALITY, 85, cv2.IMWRITE_JPEG_OPTIMIZE, 0]
            ret, buffer = cv2.imencode('.jpg', frame, encode_params)
            if not ret:
                time.sleep(0.001)  # 1ms delay (faster retry)
                continue

            # Yield frame in multipart format
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')

            last_frame_time = time.time()

    return StreamingResponse(
        generate_frames(),
        media_type="multipart/x-mixed-replace; boundary=frame"
    )


def process_rtsp_stream(session_id: str):
    """Process RTSP stream in background thread - optimized for maximum FPS (same as Live Monitoring)"""

    session = rtsp_sessions[session_id]
    rtsp_url = session['rtsp_url']
    model_name = session['model_name']

    # ✅ SUPPRESS ALL FFMPEG/OPENCV ERRORS IN CONSOLE - COMPLETE SILENCE
    import subprocess
    original_stderr = sys.stderr
    original_stdout = sys.stdout
    devnull_file = open(os.devnull, 'w')
    sys.stderr = devnull_file
    sys.stdout = devnull_file

    # Set ALL possible FFMPEG/OpenCV log levels to completely silent
    os.environ['OPENCV_FFMPEG_LOGLEVEL'] = '-8'  # AV_LOG_QUIET = -8
    os.environ['AV_LOG_FORCE_NOCOLOR'] = '1'
    os.environ['OPENCV_LOG_LEVEL'] = 'SILENT'
    os.environ['OPENCV_VIDEOIO_DEBUG'] = '0'
    os.environ['OPENCV_VIDEOIO_PRIORITY_FFMPEG'] = '1'
    os.environ['FFMPEG_LOGLEVEL'] = 'quiet'
    os.environ['AV_LOG_LEVEL'] = '-8'
    os.environ['PYTHONWARNINGS'] = 'ignore'

    # RTSP - Optimized for H.264 with error concealment
    ffmpeg_options = (
        'rtsp_transport;tcp|'           # TCP for reliability
        'fflags;nobuffer+discardcorrupt+genpts|'  # Discard corrupt + generate timestamps
        'flags;low_delay|'              # Low delay mode
        'max_delay;0|'                  # Zero delay (maximum speed)
        'reorder_queue_size;0|'         # No reordering (faster)
        'probesize;32|'                 # Minimal probing
        'analyzeduration;0|'            # Skip analysis
        'sync;ext|'                     # External sync
        'err_detect;ignore_err|'        # CRITICAL: Ignore ALL decoding errors
        'ec;favor_inter+deblock|'       # Error concealment: favor inter-frame + deblock
        'skip_frame;noref|'             # Skip non-reference frames if needed
        'framedrop;1|'                  # Allow frame dropping for speed
        'loglevel;-8|'                  # CRITICAL: Completely silent
        'allowed_media_types;video'     # Video only
    )
    os.environ['OPENCV_FFMPEG_CAPTURE_OPTIONS'] = ffmpeg_options

    cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)

    # AGGRESSIVE buffer settings for maximum FPS + error tolerance
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # Minimal buffer (1 frame) - reduces lag

    # Additional OpenCV settings for problematic streams
    cap.set(cv2.CAP_PROP_FPS, 30)  # Request 30 FPS from camera

    # Don't force codec - let FFMPEG auto-detect
    # cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'H264'))

    if not cap.isOpened():
        session['is_active'] = False
        logger.error(f"❌ Failed to open RTSP stream: {rtsp_url}")
        return

    # Get video properties for recording
    frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = int(cap.get(cv2.CAP_PROP_FPS)) or 30

    frame_count = 0
    start_time = time.time()
    last_detection_time = 0
    detection_interval = 0.5  # Run AI detection every 0.5 seconds (same as before, but with better FFMPEG settings)

    # FPS tracking
    fps_start_time = time.time()
    fps_frame_count = 0
    last_fps_log_time = time.time()

    # Cache for detection results
    cached_detections = []

    # Create recordings directory
    recordings_dir = "recordings"
    os.makedirs(recordings_dir, exist_ok=True)

    # Video recording state and steel counting for each ROI tracker
    for tracker in session['roi_trackers']:
        tracker['video_writer'] = None
        tracker['recording_start_time'] = None
        tracker['current_video_path'] = None
        # Steel counting fields
        tracker['steel_count'] = 0  # Total count of steel entering ROI (right-to-left only)
        tracker['tracked_steels'] = {}  # {track_id: {'last_x': x, 'counted': False, 'direction': 'unknown'}}

    # Track consecutive failures to detect stream issues
    consecutive_failures = 0
    max_consecutive_failures = 100

    # Frame stability tracking to reduce flickering
    last_valid_frame = None
    frame_skip_count = 0
    max_frame_skip = 3  # Skip up to 3 consecutive bad frames before using last valid

    while session['is_active'] and not session['stop_requested']:
        # Read frame as fast as possible (no blocking)
        ret, frame = cap.read()
        if not ret:
            consecutive_failures += 1
            if consecutive_failures >= max_consecutive_failures:
                logger.error(f"❌ Too many consecutive failures ({consecutive_failures}), stopping stream")
                break
            time.sleep(0.001)  # 1ms delay on failure
            continue

        # Reset failure counter on successful read
        consecutive_failures = 0

        # CRITICAL: Validate frame to skip corrupt frames
        if frame is None or frame.size == 0:
            frame_skip_count += 1
            if frame_skip_count > max_frame_skip and last_valid_frame is not None:
                frame = last_valid_frame.copy()  # Use last valid frame to prevent flickering
                frame_skip_count = 0
            else:
                continue

        # Validate frame quality - skip corrupt frames silently
        # This handles H.264 macroblock errors gracefully
        mean_brightness = np.mean(frame)
        if mean_brightness > 250 or mean_brightness < 5:  # Too bright (white) or too dark (black)
            frame_skip_count += 1
            if frame_skip_count > max_frame_skip and last_valid_frame is not None:
                frame = last_valid_frame.copy()  # Use last valid frame to prevent flickering
                frame_skip_count = 0
            else:
                continue  # Skip silently (no log spam)

        # Frame is valid - reset skip counter and save as last valid
        frame_skip_count = 0
        last_valid_frame = frame.copy()

        frame_count += 1
        fps_frame_count += 1
        session['frame_count'] = frame_count
        current_time = time.time()
        elapsed_time = current_time - start_time

        # Calculate and log FPS every 2 seconds
        if current_time - last_fps_log_time >= 2.0:
            fps_elapsed = current_time - fps_start_time
            current_fps = fps_frame_count / fps_elapsed if fps_elapsed > 0 else 0
            logger.info(f"🎥 RTSP Stream FPS: {current_fps:.2f} | Frame Count: {frame_count}")

            # Reset FPS counters
            fps_start_time = current_time
            fps_frame_count = 0
            last_fps_log_time = current_time

        # Run AI detection only every 0.5 seconds (not every frame for maximum speed)
        if current_time - last_detection_time >= detection_interval:
            last_detection_time = current_time

            # Run steel detection with TRACKING enabled (for unique IDs and counting)
            # Use track() instead of predict() to get persistent track_id for each steel
            cached_detections = ai_service.detect_objects_with_tracking(frame, model_name, confidence=0.3)

        # Check each ROI for steel presence and count (using cached detections)
        for tracker in session['roi_trackers']:
            roi_points = np.array([[p['x'], p['y']] for p in tracker['roi']], dtype=np.int32)
            steel_detected = False

            # Track steel IDs currently in ROI (for cleanup)
            current_steels_in_roi = set()

            # Check each steel detection for ROI presence and counting
            for detection in cached_detections:
                if detection['class'] != 'steel':
                    continue

                bbox = detection['bbox']
                center_x = (bbox[0] + bbox[2]) / 2
                center_y = (bbox[1] + bbox[3]) / 2
                track_id = detection['track_id']

                # Check if center point is inside ROI polygon
                if cv2.pointPolygonTest(roi_points, (center_x, center_y), False) >= 0:
                    steel_detected = True

                    # Only count if we have a valid track_id
                    if track_id is not None:
                        current_steels_in_roi.add(track_id)

                        # Count this steel if we haven't seen it before in this ROI
                        if track_id not in tracker['tracked_steels']:
                            # First time seeing this steel in ROI - COUNT IT!
                            tracker['steel_count'] += 1
                            tracker['tracked_steels'][track_id] = {
                                'counted': True,
                                'first_seen_time': time.time()
                            }
                            logger.info(f"🔢 Machine {tracker['id'] + 1}: Steel ID:{track_id} entered ROI! Count: {tracker['steel_count']}")

            # Cleanup: Remove steels that left the ROI (not seen for this frame)
            steels_to_remove = []
            for track_id in tracker['tracked_steels'].keys():
                if track_id not in current_steels_in_roi:
                    steels_to_remove.append(track_id)

            for track_id in steels_to_remove:
                del tracker['tracked_steels'][track_id]

            # Update ROI state and timing
            if steel_detected:
                # Steel present - ROI should be green, TIMER STOPS
                if tracker['is_red']:
                    # Transition from red to green - end downtime period
                    downtime_end = elapsed_time
                    downtime_start_time = tracker['downtime_start']
                    duration = downtime_end - downtime_start_time

                    # Get IST timezone
                    ist = pytz.timezone('Asia/Kolkata')
                    now = datetime.now(ist)

                    # Calculate IST timestamps for start and end
                    start_ist = now - timedelta(seconds=duration)
                    end_ist = now

                    # Stop video recording if active
                    video_path = None
                    if tracker['video_writer'] is not None:
                        tracker['video_writer'].release()
                        tracker['video_writer'] = None
                        video_path = tracker['current_video_path']
                        logger.info(f"📹 Machine {tracker['id'] + 1}: Recording stopped - {video_path}")

                    tracker['downtime_periods'].append({
                        'start': downtime_start_time,
                        'end': downtime_end,
                        'duration': duration,
                        'start_time_ist': start_ist.strftime('%H:%M:%S'),
                        'end_time_ist': end_ist.strftime('%H:%M:%S'),
                        'start_datetime_ist': start_ist.strftime('%Y-%m-%d %H:%M:%S'),
                        'end_datetime_ist': end_ist.strftime('%Y-%m-%d %H:%M:%S'),
                        'video_path': video_path if video_path else ''
                    })
                    tracker['total_downtime'] += duration

                    # Console output: Timer stopped
                    logger.info(f"⏱️  Machine {tracker['id'] + 1}: TIMER STOPPED - Steel detected! Downtime: {duration:.2f}s ({start_ist.strftime('%H:%M:%S')} to {end_ist.strftime('%H:%M:%S')})")
                tracker['is_red'] = False
            else:
                # No steel - ROI should be red, TIMER STARTS
                if not tracker['is_red']:
                    # Transition from green to red - start downtime period
                    tracker['downtime_start'] = elapsed_time
                    tracker['recording_start_time'] = current_time

                    # Console output: Timer started
                    logger.info(f"⏱️  Machine {tracker['id'] + 1}: TIMER STARTED - No steel detected at {elapsed_time:.2f}s")

                    # Start recording IMMEDIATELY when downtime starts (no 5-second delay)
                    ist = pytz.timezone('Asia/Kolkata')
                    timestamp = datetime.now(ist).strftime('%Y%m%d_%H%M%S')
                    video_filename = f"machine_{tracker['id'] + 1}_{session_id[:8]}_{timestamp}.mp4"
                    video_path = os.path.join(recordings_dir, video_filename)

                    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
                    tracker['video_writer'] = cv2.VideoWriter(video_path, fourcc, fps, (frame_width, frame_height))
                    tracker['current_video_path'] = video_path

                    logger.info(f"📹 Machine {tracker['id'] + 1}: Recording started IMMEDIATELY - {video_path}")

                # Write frame to video if recording
                if tracker['video_writer'] is not None:
                    tracker['video_writer'].write(frame)

                tracker['is_red'] = True

        # Draw ROIs, timer, and steel bounding boxes on frame (optimized)
        display_frame = draw_rois_and_timer(frame, session['roi_trackers'], frame_count, 30, cached_detections)

        # Store latest frame (atomic update) - use direct assignment for speed
        session['latest_frame'] = display_frame

        # NO SLEEP - Capture frames as fast as possible for maximum FPS
        # The frame rate is naturally limited by RTSP stream speed and AI detection interval

    # Release video capture
    cap.release()

    # Release all video writers
    for tracker in session['roi_trackers']:
        if tracker.get('video_writer') is not None:
            tracker['video_writer'].release()
            logger.info(f"📹 Machine {tracker['id'] + 1}: Recording released on session end")

    # ✅ RESTORE STDERR/STDOUT after streaming ends
    sys.stderr = original_stderr
    sys.stdout = original_stdout
    devnull_file.close()

